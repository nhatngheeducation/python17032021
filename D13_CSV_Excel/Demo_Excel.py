# pip install openpyxl: Excel > 2010, xlsx, xlsm
from openpyxl import Workbook

# Create new workbook
wb = Workbook()

# grab the active worksheet
# ws = wb.active
ws = wb.create_sheet("INFO", 1)
ws["A1"] = "Trần Văn Tèo"
ws.append([2,3,4])
# Save the file
wb.save("ExcelData.xlsx")