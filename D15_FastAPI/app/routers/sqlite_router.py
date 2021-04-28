from fastapi import APIRouter
from app.services.sqlite_service import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers
from fastapi.responses import FileResponse

router = APIRouter()

@router.get("/export_excel")
def export_hanghoa_to_excel():
    conn = sql_connection(r"data\QLHangHoa.db")
    # Lay data
    sql = """
    SELECT MaHH, TenHH, DonGia, SKU, TenLoai 
    FROM HangHoa hh JOIN Loai lo ON lo.MaLoai = hh.MaLoai
    """
    result = query_data_by_statement(conn, sql)
    conn.close()

    # Xuat data xuống file excel
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Mã hàng hóa"
    sheet.cell(row=1, column=2, value="Tên hàng hóa")
    sheet.cell(row=1, column=3, value="Đơn giá")
    sheet.cell(row=1, column=4, value="SKU")
    sheet.cell(row=1, column=5, value="Loại")
    rowIndex = 2
    for item in result:
        sheet.cell(row=rowIndex, column=1, value=item[0])
        sheet.cell(row=rowIndex, column=2, value=item[1])
        sheet.cell(row=rowIndex, column=3, value=item[2])
        sheet.cell(row=rowIndex, column=3).number_format = numbers.FORMAT_TEXT
        sheet.cell(row=rowIndex, column=4, value=item[3])
        sheet.cell(row=rowIndex, column=5, value=item[4])
        rowIndex = rowIndex + 1

    excel_file_name = "hanghoa.xlsx"
    workbook.save(filename=excel_file_name)
    workbook.close()
    # Cho phep download file
    return FileResponse(path=excel_file_name,
        filename=excel_file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/read_excel")
def read_excel_file():
    excel_file_name="ImportData.xlsx"
    wb = load_workbook(excel_file_name, read_only=True)
    sheet = wb.get_sheet_by_name('Sheet')
    rows = tuple(sheet.rows)
    row_count = len(rows)
    dshanghoa = []
    for row in rows:
        hanghoa = []
        for cell in row:
            hanghoa.append(cell.value)
        dshanghoa.append(hanghoa)

    print(dshanghoa)